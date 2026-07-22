"""
Inventory Toolkit

Stock Processing Module

Author: Gonzalo

License: MIT

Repository:
https://github.com/FPSensor/Inventory-Toolkit

"""

VERSION = "1.0.0"
import pandas as pd
import numpy as np
import argparse
import sys
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# =========================================================================
# 0. CONFIGURACIÓN CENTRALIZADA
# =========================================================================
LOCALES = ["LIBERT.R", "PASO.R", "P.OESTE", "VIRREYES"]
DEPOSITOS = {
    "LIBERT.R": "DSLIBERT",
    "PASO.R": "DSPASO.R",
    "P.OESTE": "DSPOESTE"
}

REGLAS_FAMILIAS = [
    ("56005", "Bolsas"), ("0085", "Buzos C Capucha"), ("001", "Remeras"),
    ("002", "Remeras"), ("003", "Remeras"), ("004", "Buzo Cuello redondo"),
    ("0045", "Buzo C Capucha rustico"), ("0055", "Campera Capucha rustico"),
    ("123", "Bermudas"), ("223", "Bermudas"), ("145", "Joggin Corto"),
    ("245", "Joggin Corto"), ("185", "Buzos C Capucha"), ("195", "Camperas Capucha"),
    ("008", "Buzos"), ("085", "Buzos C Capucha"), ("095", "Camperas Capucha"),
    ("099", "Camperas Capucha"), ("015", "Musculosas"), ("025", "Musculosas"),
    ("01", "Remeras"), ("O085", "Buzos C Capucha Orangutan"),
    ("O08", "Buzos C Redondo Orangutan"), ("02", "Remeras"),
    ("O03", "Camisetas Manga Larga Orangutan"), ("03", "Camisetas Manga Larga"),
    ("04", "Camisetas Manga Larga"), ("055", "Chombas Manga Larga"),
    ("05", "Chombas"), ("065", "Chombas Manga Larga"), ("06", "Chombas"),
    ("07", "Buzos medio Cierre"), ("08", "Buzo Cuello Redondo"),
    ("09", "Campera S Capucha"), ("10", "Camperas de abrigo"), ("11", "sweaters"),
    ("12", "Trajes de baño"), ("13", "Jeans"), ("14", "Joggin"),
    ("15", "Boxer"), ("16", "Buzos medio cierre"), ("17", "Camisas"),
    ("18", "Buzo Cuello redondo"), ("19", "Campera S Capucha"),
    ("20", "Camperas de Abrigo"), ("21", "sweaters"), ("22", "Trajes de baño"),
    ("23", "Jeans"), ("24", "Joggin"), ("27", "Camisas"),
    ("30", "Accesorios"), ("40", "Accesorios"), ("41", "Accesorios"),
    ("80", "Accesorios"), ("C", "Calzado"), ("S", "Jeans De segunda")
]

def clasificar_familia(codigo):
    codigo = str(codigo).strip()
    for prefijo, familia in REGLAS_FAMILIAS:
        if codigo.startswith(prefijo):
            return familia
    return "Otro"

def procesar_precios(ruta_archivo):
    try:
        df = pd.read_excel(ruta_archivo)
        df['Artículo'] = df['Artículo'].astype(str).str.split(' ').str[0]
        columnas_necesarias = ['Artículo', 'Origen - Base de datos', 'Precio']
        for col in columnas_necesarias:
            if col not in df.columns:
                print(f"Advertencia: Falta la columna '{col}' en {ruta_archivo}.")
                return None
        df = df[columnas_necesarias].rename(columns={'Origen - Base de datos': 'Base'})
        df_pivot = pd.pivot_table(
            df, index='Artículo', columns='Base', values='Precio', aggfunc='mean'
        ).reset_index()
        return df_pivot
    except Exception as e:
        print(f"Error procesando {ruta_archivo}: {e}")
        return None

def calcular_margen(df, col_venta, col_costo):
    return np.where(df[col_venta] > 0, (df[col_venta] - df[col_costo]) / df[col_venta], 0)

def dar_formato_excel(ws, is_summary=False):
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            col_name = str(ws[f"{col_letter}1"].value).upper()
            if cell.row > 1 and cell.value is not None:
                if "MARGEN" in col_name:
                    cell.number_format = '0.00%'
                elif "COSTO" in col_name or "VENTA" in col_name or "TOTAL" in col_name:
                    cell.number_format = '#,##0.00'
                elif isinstance(cell.value, (int, float)) and not is_summary:
                    cell.number_format = '#,##0'
        ws.column_dimensions[col_letter].width = (max_length + 2)

def procesar_stock(args):
    print("Iniciando procesamiento de Stock...")
    df_stock = pd.read_excel(args.stock)
    col_innecesarias = ["Grupo", "Artículo Descripción", "Color Descripción", "Talle Descripción", "CENTRAL", "DSELLERS", "SELLERS1", "DETROIT", "DSLOMAS"]
    df_stock = df_stock.drop(columns=[col for col in col_innecesarias if col in df_stock.columns], errors='ignore')
    
    for col in ['Artículo', 'Color', 'Talle']:
        df_stock[col] = df_stock[col].astype(str).str.strip()
    
    columnas_stock = ["DSLIBERT", "DSPASO.R", "DSPOESTE", "LIBERT.R", "P.OESTE", "PASO.R", "VIRREYES"]
    for col in columnas_stock:
        if col in df_stock.columns:
            df_stock[col] = (
                df_stock[col]
                .astype(str)
                .str.replace(',', '.', regex=False)
                .str.strip()
            )
            df_stock[col] = pd.to_numeric(df_stock[col], errors='coerce').fillna(0).astype(int)
    
    for local, deposito in DEPOSITOS.items():
        if local in df_stock.columns and deposito in df_stock.columns:
            df_stock[local] = df_stock[local] + df_stock[deposito]
            df_stock = df_stock.drop(columns=[deposito])
            
    df_stock['Nrw'] = df_stock.get('LIBERT.R', 0) + df_stock.get('PASO.R', 0) + df_stock.get('P.OESTE', 0)
    df_stock['Familias'] = df_stock['Artículo'].apply(clasificar_familia)
    
    print("Procesando archivos de precios...")
    df_costo = procesar_precios(args.costo)
    df_venta = procesar_precios(args.venta)
    
    # Renombramos ANTES del merge para evitar el choque de columnas (_x, _y)
    if df_costo is not None:
        renombres_costo = {col: f"PrecioUnit.Costo.{col}" for col in df_costo.columns if col != 'Artículo'}
        df_costo = df_costo.rename(columns=renombres_costo)
        df_stock = pd.merge(df_stock, df_costo, on='Artículo', how='left')
        
    if df_venta is not None:
        renombres_venta = {col: f"PrecioUnit.Venta.{col}" for col in df_venta.columns if col != 'Artículo'}
        df_venta = df_venta.rename(columns=renombres_venta)
        df_stock = pd.merge(df_stock, df_venta, on='Artículo', how='left')

    cols_precios = [c for c in df_stock.columns if c.startswith('PrecioUnit.')]
    df_stock[cols_precios] = df_stock[cols_precios].fillna(-1).astype(float)
    
    for local in LOCALES:
        col_costo, col_venta = f"PrecioUnit.Costo.{local}", f"PrecioUnit.Venta.{local}"
        if local in df_stock.columns:
            df_stock[f"{local}.Costo"] = np.where(df_stock.get(col_costo, 0) <= 0, 0, df_stock[col_costo] * df_stock[local])
            df_stock[f"{local}.Venta"] = np.where(df_stock.get(col_venta, 0) <= 0, 0, df_stock[col_venta] * df_stock[local])

    df_stock['Nrw.Costo'] = df_stock.get('LIBERT.R.Costo', 0) + df_stock.get('PASO.R.Costo', 0) + df_stock.get('P.OESTE.Costo', 0)
    df_stock['Nrw.Venta'] = df_stock.get('LIBERT.R.Venta', 0) + df_stock.get('PASO.R.Venta', 0) + df_stock.get('P.OESTE.Venta', 0)
    
    df_stock = df_stock.drop(columns=cols_precios, errors='ignore')
    
    columnas_orden = ["Artículo", "Color", "Talle", "Familias"]
    for local in LOCALES:
        columnas_orden.extend([local, f"{local}.Costo", f"{local}.Venta"])
    columnas_orden.extend(["Nrw", "Nrw.Costo", "Nrw.Venta"])
    
    df_stock = df_stock[[col for col in columnas_orden if col in df_stock.columns]]

    print("Generando reportes y aplicando formato...")
    wb = Workbook()
    wb.remove(wb.active)
    
    if all(c in df_stock.columns for c in ['LIBERT.R', 'PASO.R', 'P.OESTE', 'Nrw']):
        ws_nrw = wb.create_sheet("Nrw.I")
        df_resumen_nrw = df_stock.groupby('Familias').agg({
            'LIBERT.R': 'sum', 'LIBERT.R.Costo': 'sum', 'LIBERT.R.Venta': 'sum',
            'PASO.R': 'sum', 'PASO.R.Costo': 'sum', 'PASO.R.Venta': 'sum',
            'P.OESTE': 'sum', 'P.OESTE.Costo': 'sum', 'P.OESTE.Venta': 'sum',
            'Nrw': 'sum', 'Nrw.Costo': 'sum', 'Nrw.Venta': 'sum'
        }).reset_index()

        df_resumen_nrw['Margen_Libertad'] = calcular_margen(df_resumen_nrw, 'LIBERT.R.Venta', 'LIBERT.R.Costo')
        df_resumen_nrw['Margen_Paso'] = calcular_margen(df_resumen_nrw, 'PASO.R.Venta', 'PASO.R.Costo')
        df_resumen_nrw['Margen_Plaza'] = calcular_margen(df_resumen_nrw, 'P.OESTE.Venta', 'P.OESTE.Costo')
        df_resumen_nrw['Margen_Consolidado'] = calcular_margen(df_resumen_nrw, 'Nrw.Venta', 'Nrw.Costo')

        cols_nrw = [
            'Familias', 
            'LIBERT.R', 'LIBERT.R.Costo', 'LIBERT.R.Venta', 'Margen_Libertad',
            'PASO.R', 'PASO.R.Costo', 'PASO.R.Venta', 'Margen_Paso',
            'P.OESTE', 'P.OESTE.Costo', 'P.OESTE.Venta', 'Margen_Plaza',
            'Nrw', 'Nrw.Costo', 'Nrw.Venta', 'Margen_Consolidado'
        ]
        
        nombres_amigables_nrw = [
            'Familias', 
            'Libertad', 'CTOTAL', 'VTOTAL', 'Margen',
            'Paso', 'CTOTAL ', 'VTOTAL ', 'Margen ',
            'Plaza', 'CTOTAL  ', 'VTOTAL  ', 'Margen  ',
            'Consolidado', 'CTOTAL   ', 'VTOTAL   ', 'Margen   '
        ]
        
        df_resumen_nrw = df_resumen_nrw[cols_nrw]
        df_resumen_nrw.columns = nombres_amigables_nrw
        
        for r in dataframe_to_rows(df_resumen_nrw, index=False, header=True):
            ws_nrw.append(r)
        dar_formato_excel(ws_nrw, is_summary=True)

    if 'VIRREYES' in df_stock.columns:
        ws_vir = wb.create_sheet("Virreyes.I")
        df_resumen_vir = df_stock.groupby('Familias').agg({
            'VIRREYES': 'sum', 'VIRREYES.Costo': 'sum', 'VIRREYES.Venta': 'sum'
        }).reset_index()
        
        df_resumen_vir['Margen'] = calcular_margen(df_resumen_vir, 'VIRREYES.Venta', 'VIRREYES.Costo')
        df_resumen_vir.columns = ['Familia', 'Cantidad', 'CTOTAL', 'VTOTAL', 'Margen']
        
        for r in dataframe_to_rows(df_resumen_vir, index=False, header=True):
            ws_vir.append(r)
        dar_formato_excel(ws_vir, is_summary=True)

    ws_datos = wb.create_sheet("Datos")
    for r in dataframe_to_rows(df_stock, index=False, header=True):
        ws_datos.append(r)
    dar_formato_excel(ws_datos, is_summary=False)

    wb.save(args.out)
    print(f"¡Proceso finalizado con éxito! Archivo guardado en: {args.out}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Procesamiento y valorización de Stock.")
    parser.add_argument('-stock', required=True, help="Ruta al archivo Stock.xlsx")
    parser.add_argument('-costo', required=True, help="Ruta al archivo Costo.xlsx")
    parser.add_argument('-venta', required=True, help="Ruta al archivo Venta.xlsx")
    parser.add_argument('-out', default="Stock_Final_Reporte.xlsx", help="Ruta de salida")
    
    args = parser.parse_args()
    procesar_stock(args)
