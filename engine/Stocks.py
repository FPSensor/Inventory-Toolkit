"""
Inventory Toolkit
Stock Processing Module (Fully Abstracted & Profile-Ready)
Author: Gonzalo
License: MIT
"""

VERSION = "1.1.0"
import pandas as pd
import numpy as np
import argparse
import sys
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Importamos nuestro nuevo servicio de configuración
from core.configuration_manager import ConfigurationManager

# =========================================================================
# CORE FUNCTIONS
# =========================================================================
def clasificar_familia(codigo, familias_dict):
    """Assigns a family category based on prefix matching from external config."""
    codigo = str(codigo).strip().upper()
    
    reglas = []
    for familia, prefijos in familias_dict.items():
        for prefijo in prefijos:
            reglas.append((str(prefijo).strip().upper(), familia))
            
    reglas.sort(key=lambda x: len(x[0]), reverse=True)
    
    for prefijo, familia in reglas:
        if codigo.startswith(prefijo):
            return familia
            
    return "Otro"

def procesar_precios(ruta_archivo, pricing_dict=None):
    """Reads and processes pricing lists using fully dynamic column detection.
       Adapts to raw exports without requiring strict manual filtering."""
    if not ruta_archivo or not os.path.exists(ruta_archivo):
        return None
        
    try:
        df = pd.read_excel(ruta_archivo)
        
        # 1. Identificar columna de Artículo (Soporta nombres comunes)
        col_articulo = next((c for c in ['Artículo', 'Articulo', 'SKU', 'Codigo', 'Item'] if c in df.columns), None)
        if not col_articulo:
            print(f"WARNING: No se encontró la columna de Artículo en {ruta_archivo}.")
            return None
            
        # Limpiamos el texto del artículo (ej: "00100-151 TEE" -> "00100-151")
        df['Artículo'] = df[col_articulo].astype(str).str.split(' ').str[0]
        
        # 2. Identificar columna de Base/Sucursal
        col_base_esperada = None
        if pricing_dict and "mapeo_nombres" in pricing_dict:
            claves_dict = list(pricing_dict["mapeo_nombres"].keys())
            col_base_esperada = next((c for c in claves_dict if c in df.columns), None)

        if col_base_esperada:
            col_base = col_base_esperada
        else:
            # Fallback inteligente a columnas comunes de bases de datos
            opciones_base = ['Origen - Base de datos', 'Sucursal', 'Base', 'Local', 'Origen', 'Tienda']
            col_base = next((c for c in opciones_base if c in df.columns), None)
            
        if not col_base:
            col_base = 'Base_General'
            df[col_base] = 'General'
            
        # 3. Identificar columna de Precio/Valores
        col_valor = next((c for c in ['Precio', 'Costo', 'Venta', 'Valor', 'Monto'] if c in df.columns), None)
        
        if not col_valor:
            # Si no hay nombres obvios, agarra la última columna numérica disponible que no sea un talle
            cols_numericas = df.select_dtypes(include='number').columns.tolist()
            cols_disp = [c for c in cols_numericas if c not in [col_articulo, col_base, 'Talle', 'Color']]
            if cols_disp:
                col_valor = cols_disp[-1]
            else:
                print(f"WARNING: No se encontró una columna de valores en {ruta_archivo}.")
                return None

        # Hacemos la tabla dinámica para agrupar por sucursal y artículo sin importar el orden
        df_pivot = pd.pivot_table(
            df, 
            index='Artículo', 
            columns=col_base, 
            values=col_valor, 
            aggfunc='mean'
        ).reset_index()
        
        # Limpiamos el nombre del índice
        df_pivot.columns.name = None
        
        return df_pivot
        
    except Exception as e:
        print(f"ERROR processing {ruta_archivo}: {e}")
        return None
def calcular_margen(df, col_venta, col_costo):
    """Calculates profit margin dynamically avoiding division by zero."""
    return np.where(df[col_venta] > 0, (df[col_venta] - df[col_costo]) / df[col_venta], 0)

def dar_formato_excel(ws, is_summary=False):
    """Applies styling, bold headers, and number formatting to the Excel sheet."""
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            col_name = str(ws[f"{col_letter}1"].value).upper()
            if cell.row > 1 and cell.value is not None:
                if "MARGEN" in col_name:
                    cell.number_format = '0.00%'
                elif "COSTO" in col_name or "VENTA" in col_name or "TOTAL" in col_name:
                    cell.number_format = '#,##0.00'
                elif isinstance(cell.value, (int, float)) and not is_summary:
                    cell.number_format = '#,##0'
        ws.column_dimensions[col_letter].width = (max_length + 2)

# =========================================================================
# MAIN EXECUTION
# =========================================================================
def procesar_stock(args):
    archivos_requeridos = [args.stock, args.costo, args.venta]
    for arc in archivos_requeridos:
        if not os.path.exists(arc):
            print(f"ERROR: Data file '{arc}' not found in the current directory.")
            sys.exit(1)

    print(f"Loading external configurations for profile: {args.profile}...")
    
    # Instanciamos el manager centralizado
    config = ConfigurationManager(args.profile)

    # Asignación usando los nuevos métodos
    familias_dict = config.get_familias()
    databases_dict = config.get_databases()
    settings_dict = config.get_settings()
    stores_dict = config.get_stores()
    cleaning_dict = config.get_cleaning_rules()
    reports_dict = config.get_reports()
    
    pricing_dict = settings_dict.get("pricing", {})
    
    locales_list = stores_dict.get("locales_activos", [])
    grupos_regionales = stores_dict.get("grupos_regionales", {})
    
    unnecessary_cols = cleaning_dict.get("columnas_a_eliminar", [])
    text_cols_to_clean = cleaning_dict.get("columnas_texto_a_limpiar", ["Artículo"])
    stock_cols_to_clean = cleaning_dict.get("columnas_a_formatear", [])
    
    orden_columnas_base = reports_dict.get("orden_columnas_base", ["Artículo", "Familias"])
    datos_sheet_name = reports_dict.get("hoja_datos_crudos", "Datos")
    resumenes = reports_dict.get("resumenes", [])

    print("Starting Stock processing...")
    df_stock = pd.read_excel(args.stock)

    df_stock = df_stock.drop(columns=[col for col in unnecessary_cols if col in df_stock.columns], errors='ignore')
    
    for col in text_cols_to_clean:
        if col in df_stock.columns:
            df_stock[col] = df_stock[col].astype(str).str.strip()
    
    for col in stock_cols_to_clean:
        if col in df_stock.columns:
            df_stock[col] = (
                df_stock[col]
                .astype(str)
                .str.replace(',', '.', regex=False)
                .str.strip()
            )
            df_stock[col] = pd.to_numeric(df_stock[col], errors='coerce').fillna(0).astype(int)
    
    for local, deposito in databases_dict.items():
        if local in df_stock.columns and deposito in df_stock.columns:
            df_stock[local] = df_stock[local] + df_stock[deposito]
            df_stock = df_stock.drop(columns=[deposito])
            
    for grupo_nombre, sucursales in grupos_regionales.items():
        df_stock[grupo_nombre] = sum(df_stock.get(loc, 0) for loc in sucursales)
        
    df_stock['Familias'] = df_stock['Artículo'].apply(lambda x: clasificar_familia(x, familias_dict))
    
    print("Processing pricing files...")
    df_costo = procesar_precios(args.costo, pricing_dict)
    df_venta = procesar_precios(args.venta, pricing_dict)
    
    if df_costo is not None:
        renombres_costo = {col: f"PrecioUnit.Costo.{col}" for col in df_costo.columns if col != 'Artículo'}
        df_costo = df_costo.rename(columns=renombres_costo)
        df_stock = pd.merge(df_stock, df_costo, on='Artículo', how='left')
        
    if df_venta is not None:
        renombres_venta = {col: f"PrecioUnit.Venta.{col}" for col in df_venta.columns if col != 'Artículo'}
        df_venta = df_venta.rename(columns=renombres_venta)
        df_stock = pd.merge(df_stock, df_venta, on='Artículo', how='left')

    cols_precios = [c for c in df_stock.columns if c.startswith('PrecioUnit.')]
    df_stock[cols_precios] = df_stock[cols_precios].fillna(-1).astype(float)
    
    entidades_a_valorizar = locales_list + list(grupos_regionales.keys())
    
    for entidad in entidades_a_valorizar:
        if entidad in grupos_regionales:
            sucursales = grupos_regionales[entidad]
            df_stock[f"{entidad}.Costo"] = sum(df_stock.get(f"{loc}.Costo", 0) for loc in sucursales)
            df_stock[f"{entidad}.Venta"] = sum(df_stock.get(f"{loc}.Venta", 0) for loc in sucursales)
        else:
            col_costo = f"PrecioUnit.Costo.{entidad}"
            col_venta = f"PrecioUnit.Venta.{entidad}"
            if entidad in df_stock.columns:
                df_stock[f"{entidad}.Costo"] = np.where(df_stock.get(col_costo, 0) <= 0, 0, df_stock[col_costo] * df_stock[entidad])
                df_stock[f"{entidad}.Venta"] = np.where(df_stock.get(col_venta, 0) <= 0, 0, df_stock[col_venta] * df_stock[entidad])

    df_stock = df_stock.drop(columns=cols_precios, errors='ignore')
    
    columnas_orden = orden_columnas_base.copy()
    for entidad in entidades_a_valorizar:
        columnas_orden.extend([entidad, f"{entidad}.Costo", f"{entidad}.Venta"])
        
    df_stock = df_stock[[col for col in columnas_orden if col in df_stock.columns]]

    print("Generating dynamic reports and applying formats...")
    wb = Workbook()
    wb.remove(wb.active)
    
    for resumen in resumenes:
        nombre_hoja = resumen.get("nombre_hoja", "Resumen")
        locales_incluidos = resumen.get("locales_a_incluir", [])
        titulos_finales = resumen.get("titulos", [])
        
        entidades_validas = [loc for loc in locales_incluidos if loc in df_stock.columns]
        
        if not entidades_validas:
            continue
            
        ws = wb.create_sheet(nombre_hoja)
        
        agg_dict = {}
        for entidad in entidades_validas:
            agg_dict[entidad] = 'sum'
            if f"{entidad}.Costo" in df_stock.columns:
                agg_dict[f"{entidad}.Costo"] = 'sum'
            if f"{entidad}.Venta" in df_stock.columns:
                agg_dict[f"{entidad}.Venta"] = 'sum'
                
        df_resumen = df_stock.groupby('Familias').agg(agg_dict).reset_index()

        cols_export = ['Familias']
        for entidad in entidades_validas:
            col_costo = f"{entidad}.Costo"
            col_venta = f"{entidad}.Venta"
            col_margen = f"Margen_{entidad}"
            
            if col_venta in df_resumen.columns and col_costo in df_resumen.columns:
                df_resumen[col_margen] = calcular_margen(df_resumen, col_venta, col_costo)
                cols_export.extend([entidad, col_costo, col_venta, col_margen])
            else:
                cols_export.append(entidad)
                
        df_resumen = df_resumen[cols_export]
        
        if titulos_finales and len(titulos_finales) == len(df_resumen.columns):
            df_resumen.columns = titulos_finales
        else:
            print(f"WARNING: Títulos mismatch en hoja {nombre_hoja}. Se usarán nombres originales.")
            
        for r in dataframe_to_rows(df_resumen, index=False, header=True):
            ws.append(r)
        dar_formato_excel(ws, is_summary=True)

    ws_datos = wb.create_sheet(datos_sheet_name)
    for r in dataframe_to_rows(df_stock, index=False, header=True):
        ws_datos.append(r)
    dar_formato_excel(ws_datos, is_summary=False)

    wb.save(args.out)
    print(f"SUCCESS: Process completed! File saved at: {args.out}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Agnostic Stock Processing and Valuation Tool.")
    parser.add_argument('-profile', required=True, help="Nombre del perfil activo (ej: demo)")
    parser.add_argument('-stock', required=True, help="Path to Stock.xlsx file")
    parser.add_argument('-costo', required=True, help="Path to Costo.xlsx file")
    parser.add_argument('-venta', required=True, help="Path to Venta.xlsx file")
    parser.add_argument('-out', default="Stock_Final_Reporte.xlsx", help="Output file path")
    
    args = parser.parse_args()
    procesar_stock(args)