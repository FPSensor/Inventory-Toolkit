from core.logger import log
from core.system_utils import safe_openpyxl_save
from core.system_utils import safe_openpyxl_save

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Border, Side, Alignment
except ImportError:
    log.error("Missing openpyxl library for Excel formatting.")
    import sys
    sys.exit(1)

def apply_excel_formatting(output_file):
    wb = load_workbook(output_file)
    ws = wb.active
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    centered = Alignment(horizontal='center', vertical='center')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered
            if cell.row > 1 and isinstance(cell.value, (int, float)):
                col_name = ws.cell(row=1, column=cell.column).value
                if col_name in ['Diferencia', 'CTOTAL', 'VTOTAL']:
                    if cell.value < 0: cell.fill = red_fill
                    elif cell.value > 0: cell.fill = green_fill
                if col_name in ['CTOTAL', 'VTOTAL']:
                    cell.number_format = '#,##0.00'
    return safe_openpyxl_save(wb, output_file)
