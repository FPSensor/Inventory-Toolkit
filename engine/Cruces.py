"""
Inventory Toolkit
Inventory Processing Module (Cruces)
Author: Gonzalo
License: MIT
"""
VERSION = "1.5.0 (CLI Integration)"
import pandas as pd
import sys
import os

from core.configuration_manager import ConfigurationManager 

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Border, Side, Alignment
except ImportError:
    print("ERROR: Falta la librería openpyxl para darle formato al Excel.")
    sys.exit(1)

def construir_reglas_familia(familias_dict):
    reglas = []
    for familia, prefijos in familias_dict.items():
        for prefijo in prefijos:
            reglas.append((str(prefijo).strip().upper(), familia))
            
    # Ordena dinámicamente de mayor a menor longitud del prefijo
    reglas.sort(key=lambda x: len(x[0]), reverse=True)
    return reglas

def asignar_familia(codigo, reglas):
    if pd.isna(codigo) or not isinstance(codigo, str): 
        return "Otro"
        
    codigo = str(codigo).strip().upper()
    if codigo.startswith("REVISAR"): 
        return "REVISAR"
    
    for prefijo, familia in reglas:
        if codigo.startswith(prefijo): 
            return familia
    return "Otro"

def normalizar_articulo(lectura, master_base):
    if pd.isna(lectura):
        return None
    lectura_upper = str(lectura).upper().strip()
    coincidencias = [art for art in master_base if lectura_upper.startswith(str(art).upper())]
    
    if not coincidencias:
        return f"REVISAR | {lectura}"
        
    coincidencias.sort(key=len, reverse=True)
    return coincidencias[0]

def calcular_diferencia(stock, conteo):
    if stock < 0 and conteo > 0:
        return conteo
    return conteo - stock

def aplicar_formato_excel(archivo_salida):
    wb = load_workbook(archivo_salida)
    ws = wb.active

    borde_fino = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    centrado = Alignment(horizontal='center', vertical='center')
    fondo_verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fondo_rojo = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = borde_fino
            cell.alignment = centrado
            
            if cell.row > 1 and isinstance(cell.value, (int, float)):
                col_name = ws.cell(row=1, column=cell.column).value
                if col_name in ['Diferencia', 'CTOTAL', 'VTOTAL']:
                    if cell.value < 0:
                        cell.fill = fondo_rojo
                    elif cell.value > 0:
                        cell.fill = fondo_verde
                if col_name in ['CTOTAL', 'VTOTAL']:
                    cell.number_format = '#,##0.00'

    wb.save(archivo_salida)

def ejecutar_cruce(args):
    """
    Punto de entrada para el CLI. Recibe el namespace 'args' 
    con las rutas y las banderas.
    """
    archivos_requeridos = [args.sistema, args.conteo, args.costo, args.venta]
    for arc in archivos_requeridos:
        if not os.path.exists(arc):
            print(f"ERROR: No se encontró el archivo '{arc}'.")
            return  # Usamos return en vez de sys.exit para no matar el CLI entero

    print(f"Cargando configuración para el perfil: {args.profile}...")
    config = ConfigurationManager(profile=args.profile)
    
    familias_raw = config.get_familias()
    reglas_familia = construir_reglas_familia(familias_raw)

    try:
        settings_cruces = config.cruces_settings
        art_ignorados = settings_cruces.get("articulos_ignorados", [])
        pal_ignoradas = settings_cruces.get("palabras_ignoradas", [])
        col_costo_art = settings_cruces.columnas_costo.get("articulo", "Artículo - Código")
        col_costo_precio = settings_cruces.columnas_costo.get("precio", "Precio")
        col_venta_art = settings_cruces.columnas_venta.get("articulo", "Artículo - Código")
        col_venta_precio = settings_cruces.columnas_venta.get("precio", "Precio")
    except AttributeError:
        art_ignorados = ['12060-142', '12102-142']
        pal_ignoradas = ['Total general']
        col_costo_art, col_costo_precio = 'Artículo - Código', 'Precio'
        col_venta_art, col_venta_precio = 'Artículo - Código', 'Precio'

    print("Procesando archivos...")
    df_stock = pd.read_excel(args.sistema)
    df_conteo = pd.read_excel(args.conteo, header=None, names=['Artículo_Lectura'])
    df_costo = pd.read_excel(args.costo)
    df_venta = pd.read_excel(args.venta)

    df_stock['Artículo'] = df_stock['Artículo'].astype(str).str.strip()
    df_stock['Cantidad'] = pd.to_numeric(df_stock['Cantidad'], errors='coerce').fillna(0)
    master_base = df_stock['Artículo'].unique().tolist()

    df_costo.rename(columns={col_costo_art: 'Artículo', col_costo_precio: 'Costo'}, inplace=True)
    df_venta.rename(columns={col_venta_art: 'Artículo', col_venta_precio: 'Precio'}, inplace=True)
    df_costo['Artículo'] = df_costo['Artículo'].astype(str).str.strip().apply(lambda x: x.split()[0] if x else x)
    df_venta['Artículo'] = df_venta['Artículo'].astype(str).str.strip().apply(lambda x: x.split()[0] if x else x)

    df_conteo = df_conteo.dropna(subset=['Artículo_Lectura'])
    df_conteo['Artículo_Lectura'] = df_conteo['Artículo_Lectura'].astype(str).str.strip()
    df_conteo['Total_Original'] = 1
    
    print("Normalizando códigos escaneados (esto puede tardar unos segundos)...")
    df_conteo['Artículo'] = df_conteo['Artículo_Lectura'].apply(lambda x: normalizar_articulo(x, master_base))

    if args.ds:
        df_stock_cons = df_stock.groupby('Artículo', as_index=False)['Cantidad'].sum()
    else:
        df_stock_cons = df_stock[['Artículo', 'Cantidad']].copy()
        
    df_stock_cons.rename(columns={'Cantidad': 'Stock Sistema'}, inplace=True)
    df_conteo_cons = df_conteo.groupby('Artículo', as_index=False)['Total_Original'].sum()
    df_conteo_cons.rename(columns={'Total_Original': 'Conteo Físico'}, inplace=True)

    df_costo_cons = df_costo.groupby('Artículo')['Costo'].mean().reset_index()
    df_venta_cons = df_venta.groupby('Artículo')['Precio'].mean().reset_index()

    if args.parcial:
        df_cruce = pd.merge(df_conteo_cons, df_stock_cons, on='Artículo', how='left')
    else:
        df_cruce = pd.merge(df_conteo_cons, df_stock_cons, on='Artículo', how='outer')
        
    df_cruce['Conteo Físico'] = df_cruce['Conteo Físico'].fillna(0)
    df_cruce['Stock Sistema'] = df_cruce['Stock Sistema'].fillna(0)

    if art_ignorados:
        df_cruce = df_cruce[~df_cruce['Artículo'].isin(art_ignorados)]
    for palabra in pal_ignoradas:
        df_cruce = df_cruce[~df_cruce['Artículo'].str.contains(palabra, case=False, na=False)]

    df_cruce['Diferencia'] = df_cruce.apply(lambda row: calcular_diferencia(row['Stock Sistema'], row['Conteo Físico']), axis=1)
    df_cruce['Familias'] = df_cruce['Artículo'].apply(lambda x: asignar_familia(x, reglas_familia))

    df_final = df_cruce.merge(df_costo_cons, on='Artículo', how='left')
    df_final = df_final.merge(df_venta_cons, on='Artículo', how='left')

    df_final['CTOTAL'] = df_final['Diferencia'] * df_final['Costo'].fillna(0)
    df_final['VTOTAL'] = df_final['Diferencia'] * df_final['Precio'].fillna(0)

    df_final = df_final[df_final['Diferencia'] != 0].copy()
    cols = ['Familias', 'Artículo', 'Stock Sistema', 'Conteo Físico', 'Diferencia', 'CTOTAL', 'VTOTAL']
    df_final = df_final[cols].sort_values(by=['Familias', 'Artículo'])

    df_final.to_excel(args.out, index=False)
    aplicar_formato_excel(args.out)
    
    print(f"¡Proceso terminado con éxito! Tu archivo está listo, centrado, pintado y formateado en: {args.out}")