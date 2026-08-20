from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from core.logger import log
from core.system_utils import safe_openpyxl_save
from core.system_utils import safe_openpyxl_save
from engine.stock_processing.data_processor import calculate_margin

def apply_excel_formatting(ws, is_summary=False):
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            col_name = str(ws[f"{col_letter}1"].value).upper()
            if cell.row > 1 and cell.value is not None:
                if "MARGEN" in col_name:
                    cell.number_format = '0.00%'
                elif "COSTO" in col_name or "VENTA" in col_name or "TOTAL" in col_name:
                    cell.number_format = '#,##0.00'
                elif isinstance(cell.value, (int, float)) and not is_summary:
                    cell.number_format = '#,##0'
        ws.column_dimensions[col_letter].width = (max_length + 2)

def render_stock_excel(output_file, df_stock, summaries, df_columns, raw_data_sheet):
    log.info("Generating dynamic reports and applying formats...")
    wb = Workbook()
    wb.remove(wb.active)
    
    for summary in summaries:
        sheet_name = summary.get("nombre_hoja", "Resumen")
        included_stores = summary.get("locales_a_incluir", [])
        final_titles = summary.get("titulos", [])
        
        valid_entities = [loc for loc in included_stores if loc in df_columns]
        if not valid_entities: continue
            
        ws = wb.create_sheet(sheet_name)
        agg_dict = {}
        for entity in valid_entities:
            agg_dict[entity] = 'sum'
            if f"{entity}.Costo" in df_columns: agg_dict[f"{entity}.Costo"] = 'sum'
            if f"{entity}.Venta" in df_columns: agg_dict[f"{entity}.Venta"] = 'sum'
                
        df_summary = df_stock.groupby('Familias').agg(agg_dict).reset_index()

        export_cols = ['Familias']
        for entity in valid_entities:
            col_cost = f"{entity}.Costo"
            col_sales = f"{entity}.Venta"
            col_margin = f"Margen_{entity}"
            
            if col_sales in df_summary.columns and col_cost in df_summary.columns:
                df_summary[col_margin] = calculate_margin(df_summary, col_sales, col_cost)
                export_cols.extend([entity, col_cost, col_sales, col_margin])
            else:
                export_cols.append(entity)
                
        df_summary = df_summary[export_cols]
        
        if final_titles and len(final_titles) == len(df_summary.columns):
            df_summary.columns = final_titles
        else:
            log.warning(f"Title mismatch in sheet {sheet_name}. Original names will be used.")
            
        for r in dataframe_to_rows(df_summary, index=False, header=True):
            ws.append(r)
        apply_excel_formatting(ws, is_summary=True)

    ws_data = wb.create_sheet(raw_data_sheet)
    for r in dataframe_to_rows(df_stock, index=False, header=True):
        ws_data.append(r)
    apply_excel_formatting(ws_data, is_summary=False)

    return safe_openpyxl_save(wb, output_file)
