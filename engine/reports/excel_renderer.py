import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def apply_style(cell, is_header=False, is_total=False, num_format=None):
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    if is_header:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    if is_total:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    if num_format:
        cell.number_format = num_format

def _render_sheet(ws, df_curr, df_prev, start_dt, end_dt, start_prev, config, grouping_col):
    col_s = config["data_source"]["branch_column"]
    col_c = config["data_source"]["quantity_column"]
    groups = config["report_structures"]
    
    all_branches = list(dict.fromkeys([b for sucs in groups.values() for b in sucs]))
    
    total_cols_b1 = len(all_branches) + 2
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols_b1)
    title_cell = ws['A1']
    
    if pd.isna(start_dt) or pd.isna(end_dt):
        title_cell.value = "Sales (No Data)"
    else:
        title_cell.value = f'Sales from {start_dt.strftime("%B %d, %Y")} to {end_dt.strftime("%B %d, %Y")}'
        
    apply_style(title_cell, is_header=True)

    headers_b1 = [grouping_col] + all_branches + ['Totals']
    for col_idx, h in enumerate(headers_b1, 1):
        apply_style(ws.cell(row=2, column=col_idx, value=h), is_header=True)

    t1 = pd.pivot_table(df_curr, values=col_c, index=grouping_col, columns=col_s, aggfunc='sum', fill_value=0)
    for b in all_branches:
        if b not in t1.columns: t1[b] = 0
    t1 = t1[all_branches].sort_index()

    current_row = 3
    start_row_b1 = current_row
    for item, row in t1.iterrows():
        apply_style(ws.cell(row=current_row, column=1, value=item))
        for i, b in enumerate(all_branches):
            val = row[b]
            apply_style(ws.cell(row=current_row, column=2+i, value=val if val != 0 else None), num_format='#,##0')
        
        cols_letters = [get_column_letter(2+i) for i in range(len(all_branches))]
        apply_style(ws.cell(row=current_row, column=total_cols_b1, value="=" + "+".join([f"{let}{current_row}" for let in cols_letters])), num_format='#,##0', is_total=True)
        current_row += 1

    end_row_b1 = current_row - 1
    if end_row_b1 < start_row_b1:
        end_row_b1 = start_row_b1
        
    apply_style(ws.cell(row=current_row, column=1, value='Totals'), is_total=True)
    for i in range(len(all_branches) + 1):
        let = get_column_letter(2+i)
        apply_style(ws.cell(row=current_row, column=2+i, value=f'=SUM(${let}{start_row_b1}:${let}{end_row_b1})'), is_total=True, num_format='#,##0')

    current_row += 3

    for group_name, branches in groups.items():
        if not branches:
            continue
        
        group_cols = 1 + len(branches) * 3
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=group_cols)
        apply_style(ws.cell(row=current_row, column=1, value=f'YoY Comparison - {group_name.replace("_", " ").title()}'), is_header=True)
        current_row += 1

        idx = 2
        for b in branches:
            ws.merge_cells(start_row=current_row, start_column=idx, end_row=current_row, end_column=idx+1)
            apply_style(ws.cell(row=current_row, column=idx, value=b), is_header=True)
            apply_style(ws.cell(row=current_row, column=idx+1), is_header=True)
            apply_style(ws.cell(row=current_row, column=idx+2), is_header=True)
            idx += 3
        current_row += 1

        apply_style(ws.cell(row=current_row, column=1, value=grouping_col), is_header=True)
        idx = 2
        for b in branches:
            apply_style(ws.cell(row=current_row, column=idx, value=start_prev.year), is_header=True)
            
            curr_year = end_dt.year if pd.notna(end_dt) else start_dt.year + 1
            apply_style(ws.cell(row=current_row, column=idx+1, value=curr_year), is_header=True)
            apply_style(ws.cell(row=current_row, column=idx+2, value='%'), is_header=True)
            idx += 3

        t_prev = pd.pivot_table(df_prev[df_prev[col_s].isin(branches)], values=col_c, index=grouping_col, columns=col_s, aggfunc='sum', fill_value=0)
        t_curr = pd.pivot_table(df_curr[df_curr[col_s].isin(branches)], values=col_c, index=grouping_col, columns=col_s, aggfunc='sum', fill_value=0)
        
        all_items = list(dict.fromkeys(list(t_prev.index) + list(t_curr.index)))
        all_items.sort()
        
        for b in branches:
            if b not in t_prev.columns: t_prev[b] = 0
            if b not in t_curr.columns: t_curr[b] = 0
            
        t_prev = t_prev.reindex(all_items, fill_value=0)
        t_curr = t_curr.reindex(all_items, fill_value=0)
        
        current_row += 1
        start_row_g = current_row
        for item in all_items:
            apply_style(ws.cell(row=current_row, column=1, value=item))
            idx = 2
            for b in branches:
                val_p = t_prev.loc[item, b]
                val_c = t_curr.loc[item, b]
                apply_style(ws.cell(row=current_row, column=idx, value=val_p if val_p != 0 else None), num_format='#,##0')
                apply_style(ws.cell(row=current_row, column=idx+1, value=val_c if val_c != 0 else None), num_format='#,##0')
                
                let_p = get_column_letter(idx)
                let_c = get_column_letter(idx+1)
                apply_style(ws.cell(row=current_row, column=idx+2, value=f'=IF({let_p}{current_row}=0, "N/A", {let_c}{current_row}/{let_p}{current_row}-1)'), num_format='0.00%')
                idx += 3
            current_row += 1

        end_row_g = current_row - 1
        if end_row_g < start_row_g:
            end_row_g = start_row_g
            
        apply_style(ws.cell(row=current_row, column=1, value='Totals'), is_total=True)
        idx = 2
        for b in branches:
            let_p = get_column_letter(idx)
            let_c = get_column_letter(idx+1)
            apply_style(ws.cell(row=current_row, column=idx, value=f'=SUM(${let_p}{start_row_g}:${let_p}{end_row_g})'), is_total=True, num_format='#,##0')
            apply_style(ws.cell(row=current_row, column=idx+1, value=f'=SUM(${let_c}{start_row_g}:${let_c}{end_row_g})'), is_total=True, num_format='#,##0')
            apply_style(ws.cell(row=current_row, column=idx+2, value=f'=IF({let_p}{current_row}=0, "N/A", {let_c}{current_row}/{let_p}{current_row}-1)'), is_total=True, num_format='0.00%')
            idx += 3

        current_row += 3

    max_col = ws.max_column
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['A'].width = 30


def render_yoy_sales_excel(output_path, df_curr, df_prev, start_dt, end_dt, start_prev, config, grouping_col, segmented):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    col_f = config["data_source"]["date_column"]
    df_curr_clean = df_curr.dropna(subset=[col_f]) if col_f in df_curr.columns else df_curr
    
    # 1. Pesta$as segmentadas por mes
    if segmented and not df_curr_clean.empty:
        periods = df_curr_clean[col_f].dt.to_period('M').unique()
        periods = sorted(periods)
        
        for p in periods:
            mask_c = df_curr[col_f].dt.to_period('M') == p
            df_c = df_curr[mask_c]
            
            mask_p = df_prev[col_f].dt.to_period('M') == (p - 12)
            df_p = df_prev[mask_p]
            
            sheet_name = p.strftime('%m-%y')
            ws = wb.create_sheet(title=sheet_name)
            
            s_dt = df_c[col_f].min()
            e_dt = df_c[col_f].max()
            if pd.isna(s_dt): s_dt = start_dt
            if pd.isna(e_dt): e_dt = end_dt
            s_prev = s_dt - pd.DateOffset(years=1)
            
            _render_sheet(ws, df_c, df_p, s_dt, e_dt, s_prev, config, grouping_col)
            
    # 2. La magia del reporte anual
    span_days = (end_dt - start_dt).days
    
    if span_days > 366 and not df_curr_clean.empty:
        # Detectamos que abarca m#ltiples a$os, segmentamos los totales por a$o
        years = df_curr_clean[col_f].dt.year.unique()
        years = sorted(years)
        
        for y in years:
            mask_c = df_curr[col_f].dt.year == y
            df_c = df_curr[mask_c]
            
            mask_p = df_prev[col_f].dt.year == (y - 1)
            df_p = df_prev[mask_p]
            
            sheet_name = f"Full {y}"
            ws_full = wb.create_sheet(title=sheet_name)
            
            s_dt = df_c[col_f].min()
            e_dt = df_c[col_f].max()
            if pd.isna(s_dt): s_dt = start_dt
            if pd.isna(e_dt): e_dt = end_dt
            s_prev = s_dt - pd.DateOffset(years=1)
            
            _render_sheet(ws_full, df_c, df_p, s_dt, e_dt, s_prev, config, grouping_col)
    else:
        # Lapso normal (hasta 1 a$o), armamos la hoja cl sica
        title = "Full Report" if segmented else "Sales"
        ws_full = wb.create_sheet(title=title)
        _render_sheet(ws_full, df_curr, df_prev, start_dt, end_dt, start_prev, config, grouping_col)

    # 3. Fallback por si la limpieza dej" el Excel sin pesta$as
    if not wb.sheetnames:
        ws_empty = wb.create_sheet(title="Sales")
        _render_sheet(ws_empty, df_curr, df_prev, start_dt, end_dt, start_prev, config, grouping_col)

    wb.save(output_path)
