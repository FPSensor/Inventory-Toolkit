"""
Inventory Toolkit

Inventory Processing Module

Author: Gonzalo

License: MIT

Repository:
https://github.com/FPSensor/Inventory-Toolkit

"""
VERSION = "1.0.0"
import pandas as pd
import argparse
import sys
import os
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Border, Side, Alignment
except ImportError:
    print("ERROR: Falta la librería openpyxl para darle formato al Excel.")
    print("Instalala ejecutando: pip install openpyxl")
    sys.exit(1)

def asignar_familia(codigo):
    if pd.isna(codigo) or not isinstance(codigo, str): 
        return "Otro"
        
    # Limpiamos y estandarizamos el código de entrada a mayúsculas
    codigo = str(codigo).strip().upper()
    
    if codigo.startswith("REVISAR"): 
        return "REVISAR"
    
    # Diccionario maestro estandarizado (sin duplicados por mayúsculas/minúsculas)
    reglas = [
        ("56005", "Bolsas"),
        ("0085", "Buzos C Capucha"),
        ("0045", "Buzo C Capucha Rustico"),
        ("0055", "Campera Capucha Rustico"),
        ("O085", "Buzos C Capucha Orangutan"),
        ("O03", "Camisetas Manga Larga Orangutan"),
        ("O08", "Buzos C Redondo Orangutan"),
        ("123", "Bermudas"),
        ("223", "Bermudas"),
        ("145", "Jogging Corto"),
        ("245", "Jogging Corto"),
        ("185", "Buzos C Capucha"),
        ("195", "Camperas Capucha"),
        ("085", "Buzos C Capucha"),
        ("095", "Camperas Capucha"),
        ("099", "Camperas Capucha"),
        ("015", "Musculosas"),
        ("025", "Musculosas"),
        ("055", "Chombas Manga Larga"),
        ("065", "Chombas Manga Larga"),
        ("001", "Remeras"),
        ("002", "Remeras"),
        ("003", "Remeras"),
        ("004", "Buzo Cuello Redondo"),
        ("008", "Buzos"),
        ("01", "Remeras"),
        ("02", "Remeras"),
        ("03", "Camisetas Manga Larga"),
        ("04", "Camisetas Manga Larga"),
        ("05", "Chombas"),
        ("06", "Chombas"),
        ("07", "Buzos Medio Cierre"),
        ("08", "Buzo Cuello Redondo"),
        ("09", "Campera S Capucha"),
        ("10", "Camperas De Abrigo"),
        ("11", "Sweaters"),
        ("12", "Trajes De Baño"),
        ("13", "Jeans"),
        ("14", "Jogging"),
        ("15", "Boxer"),
        ("16", "Buzos Medio Cierre"),
        ("17", "Camisas"),
        ("18", "Buzo Cuello Redondo"),
        ("19", "Campera S Capucha"),
        ("20", "Camperas De Abrigo"),
        ("21", "Sweaters"),
        ("22", "Trajes De Baño"),
        ("23", "Jeans"),
        ("24", "Jogging"),
        ("27", "Camisas"),
        ("30", "Accesorios"),
        ("40", "Accesorios"),
        ("41", "Accesorios"),
        ("80", "Accesorios"),
        ("D", "Detroit"),
        ("C", "Calzado"),
        ("S", "Jeans De Segunda")
    ]
    
    # Ordena dinámicamente de mayor a menor longitud del prefijo para evitar pisar reglas
    reglas.sort(key=lambda x: len(x[0]), reverse=True)
    
    for prefijo, familia in reglas:
        if codigo.startswith(prefijo): 
            return familia
            
    return "Otro"

def normalizar_articulo(lectura, master_base):
    if pd.isna(lectura):
        return None
    lectura_upper = str(lectura).upper().strip()
    coincidencias = [art for art in master_base if lectura_upper.startswith(str(art).upper())]
    
    if not coincidencias:
        return f"REVISAR | {lectura}"
        
    coincidencias.sort(key=len, reverse=True)
    return coincidencias[0]

def calcular_diferencia(stock, conteo):
    if stock < 0 and conteo > 0:
        return conteo
    return conteo - stock

def aplicar_formato_excel(archivo_salida):
    wb = load_workbook(archivo_salida)
    ws = wb.active

    borde_fino = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    centrado = Alignment(horizontal='center', vertical='center')
    
    fondo_verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fondo_rojo = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = borde_fino
            cell.alignment = centrado
            
            if cell.row > 1 and isinstance(cell.value, (int, float)):
                col_name = ws.cell(row=1, column=cell.column).value
                
                # Colorear diferencias
                if col_name in ['Diferencia', 'CTOTAL', 'VTOTAL']:
                    if cell.value < 0:
                        cell.fill = fondo_rojo
                    elif cell.value > 0:
                        cell.fill = fondo_verde
                        
                # Aplicar separador de miles y 2 decimales para CTOTAL y VTOTAL
                if col_name in ['CTOTAL', 'VTOTAL']:
                    cell.number_format = '#,##0.00'

    wb.save(archivo_salida)

def main():
    parser = argparse.ArgumentParser(description="Automatización de Cruce de Inventario")
    parser.add_argument("-sistema", default="Sistema.xls", help="Archivo de stock del sistema")
    parser.add_argument("-conteo", default="Conteo.xlsx", help="Archivo de conteo físico")
    parser.add_argument("-costo", default="Costo.xlsx", help="Archivo de lista de costos")
    parser.add_argument("-venta", default="Venta.xlsx", help="Archivo de lista de ventas")
    parser.add_argument("-out", default="Cruce_Final.xlsx", help="Nombre del archivo de salida")
    parser.add_argument("-ds", action="store_true", help="Consolidar cantidades de múltiples bases")
    parser.add_argument("-parcial", action="store_true", help="Filtrar stock por artículos escaneados")
    
    args = parser.parse_args()

    archivos_requeridos = [args.sistema, args.conteo, args.costo, args.venta]
    for arc in archivos_requeridos:
        if not os.path.exists(arc):
            print(f"ERROR: No se encontró el archivo '{arc}'.")
            sys.exit(1)

    df_stock = pd.read_excel(args.sistema)
    df_conteo = pd.read_excel(args.conteo, header=None, names=['Artículo_Lectura'])
    df_costo = pd.read_excel(args.costo)
    df_venta = pd.read_excel(args.venta)

    df_stock['Artículo'] = df_stock['Artículo'].astype(str).str.strip()
    df_stock['Cantidad'] = pd.to_numeric(df_stock['Cantidad'], errors='coerce').fillna(0)
    master_base = df_stock['Artículo'].unique().tolist()

    df_costo.rename(columns={'Artículo - Código': 'Artículo', 'Precio': 'Costo'}, inplace=True)
    df_venta.rename(columns={'Artículo - Código': 'Artículo', 'Precio': 'Precio'}, inplace=True)
    df_costo['Artículo'] = df_costo['Artículo'].astype(str).str.strip()
    df_venta['Artículo'] = df_venta['Artículo'].astype(str).str.strip()

    df_conteo = df_conteo.dropna(subset=['Artículo_Lectura'])
    df_conteo['Artículo_Lectura'] = df_conteo['Artículo_Lectura'].astype(str).str.strip()
    df_conteo['Total_Original'] = 1
    df_conteo['Artículo'] = df_conteo['Artículo_Lectura'].apply(lambda x: normalizar_articulo(x, master_base))

    if args.ds:
        df_stock_cons = df_stock.groupby('Artículo', as_index=False)['Cantidad'].sum()
    else:
        df_stock_cons = df_stock[['Artículo', 'Cantidad']].copy()
        
    df_stock_cons.rename(columns={'Cantidad': 'Stock Sistema'}, inplace=True)
    df_conteo_cons = df_conteo.groupby('Artículo', as_index=False)['Total_Original'].sum()
    df_conteo_cons.rename(columns={'Total_Original': 'Conteo Físico'}, inplace=True)

    df_costo_cons = df_costo.groupby('Artículo')['Costo'].mean().reset_index()
    df_venta_cons = df_venta.groupby('Artículo')['Precio'].mean().reset_index()

    if args.parcial:
        df_cruce = pd.merge(df_conteo_cons, df_stock_cons, on='Artículo', how='left')
    else:
        df_cruce = pd.merge(df_conteo_cons, df_stock_cons, on='Artículo', how='outer')
        
    df_cruce['Conteo Físico'] = df_cruce['Conteo Físico'].fillna(0)
    df_cruce['Stock Sistema'] = df_cruce['Stock Sistema'].fillna(0)

    df_cruce = df_cruce[~df_cruce['Artículo'].isin(['12060-142', '12102-142'])]
    df_cruce = df_cruce[~df_cruce['Artículo'].str.contains('Total general', case=False, na=False)]

    df_cruce['Diferencia'] = df_cruce.apply(lambda row: calcular_diferencia(row['Stock Sistema'], row['Conteo Físico']), axis=1)
    df_cruce['Familias'] = df_cruce['Artículo'].apply(asignar_familia)

    df_final = df_cruce.merge(df_costo_cons, on='Artículo', how='left')
    df_final = df_final.merge(df_venta_cons, on='Artículo', how='left')

    df_final['CTOTAL'] = df_final['Diferencia'] * df_final['Costo'].fillna(0)
    df_final['VTOTAL'] = df_final['Diferencia'] * df_final['Precio'].fillna(0)

    df_final = df_final[df_final['Diferencia'] != 0].copy()
    cols = ['Familias', 'Artículo', 'Stock Sistema', 'Conteo Físico', 'Diferencia', 'CTOTAL', 'VTOTAL']
    df_final = df_final[cols].sort_values(by=['Familias', 'Artículo'])

    df_final.to_excel(args.out, index=False)
    
    aplicar_formato_excel(args.out)
    print(f"¡Proceso terminado con éxito! Tu archivo está listo, centrado, pintado y formateado en: {args.out}")

if __name__ == "__main__":
    main()